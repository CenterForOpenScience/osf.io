import logging

import requests
from rest_framework import status as http_status

from framework.exceptions import HTTPError
from osf.models import BaseFileNode
from osf.models.files import UnableToResolveFileClass
from . import SHORT_NAME

import json

from io import BytesIO
from website import settings
from flask import request
import numpy as np
import openpyxl
import io
from PIL import Image
import imghdr
import csv
import chardet
from addons.metadata.apps import AddonAppConfig as AddonAppConfig

import mimetypes
from api.base.utils import waterbutler_api_url_for

from .suggestions import suggestion_erad, suggestion_contributor, suggest_kaken


logger = logging.getLogger(__name__)


ROR_URL = 'https://api.ror.org/v2/organizations'


def valid_suggestion_key(key):
    if key == 'file-data-number':
        return True
    elif key.startswith('get-'):
        return True
    elif key == 'ror':
        return True
    elif key.startswith('erad:'):
        return True
    elif key.startswith('kaken:'):
        return True
    elif key.startswith('asset:'):
        return True
    elif key.startswith('contributor:'):
        return True
    return False


def suggestion_metadata(key, keyword, filepath, node):
    suggestions = []
    if key == 'file-data-number':
        suggestions.extend(suggestion_file_data_number(key, filepath, node))
    elif key.startswith('get-'):
        suggestions.extend(suggestion_file_metadata_get_value(key, keyword, filepath, node))
    elif key == 'ror':
        suggestions.extend(suggestion_ror(key, keyword))
    elif key.startswith('erad:'):
        suggestions.extend(suggestion_erad(key, keyword, node))
    elif key.startswith('kaken:'):
        suggestions.extend(suggest_kaken(key, keyword, node))
    elif key.startswith('asset:'):
        suggestions.extend(suggestion_asset(key, keyword, node))
    elif key.startswith('contributor:'):
        suggestions.extend(suggestion_contributor(key, keyword, node))
    else:
        raise KeyError('Invalid key: {}'.format(key))
    return suggestions

def suggestion_file_metadata_get_value(key, keyword, filepath, node):
    data = ''
    error_string = ''
    extension = ''
    try:
        parts = filepath.split('/', 1)
        provider = parts[0]
        path = '/' + parts[1]

        cookie = request.cookies.get(settings.COOKIE_NAME)
        cookies = {settings.COOKIE_NAME: cookie}

        url = waterbutler_api_url_for(node._id, provider, path, _internal=True)

        file_size, extension = get_file_size_and_extension(url, cookies)

        excel_file_maximum_size = AddonAppConfig.excel_file_maximum_size
        text_file_maximum_size = AddonAppConfig.text_file_maximum_size
        image_file_maximum_size = AddonAppConfig.image_file_maximum_size
        any_file_maximum_size = AddonAppConfig.any_file_maximum_size

        if (key == 'get-excel-row-count' or key == 'get-excel-column-count') and extension in AddonAppConfig.excel_file_extension:
            if file_size < excel_file_maximum_size:
                response = download_file(url, cookies)
                try:
                    byteCode = BytesIO(response)
                    workbook = openpyxl.load_workbook(byteCode, data_only=True)
                    sheet_name = workbook.sheetnames[0]
                    worksheet = workbook[sheet_name]
                    excel_data = np.array([[cell.value for cell in row] for row in worksheet.iter_rows()])
                    rows, columns = excel_data.shape
                    data = str(rows) + str(columns)

                    if key == 'get-excel-row-count':
                        data = str(rows)
                    elif key == 'get-excel-column-count':
                        data = str(columns)
                except Exception as e:
                    data = ''
                    error_string = error_string + str(e)
                    logger.info(error_string)
            else:
                data = 'get-filesize-over-error'

        elif key == 'get-text-row-count' and extension in AddonAppConfig.text_file_extension:
            if file_size < text_file_maximum_size:
                response = download_file(url, cookies)
                try:
                    encoding_info = chardet.detect(response)
                    text_data = encoding_info['encoding']
                    content = response.decode(text_data)
                    csv_data = io.StringIO(content)

                    row_count = sum(1 for _ in csv_data)
                    data = str(row_count)

                except Exception as e:
                    data = ''
                    error_string = error_string + str(e)
                    logger.info(error_string)
            else:
                data = 'get-filesize-over-error'

        elif key == 'get-text-column-count' and extension in AddonAppConfig.text_file_extension:
            if file_size < text_file_maximum_size:
                response = download_file(url, cookies)
                try:
                    text_data = chardet.detect(response).get('encoding')
                    content = response.decode(text_data)

                # Detect the delimiter
                    delimiter = detect_delimiter(content, AddonAppConfig.delimiters)

                    if delimiter:
                        csv_data = io.StringIO(content)
                        csv_reader = csv.reader(csv_data, delimiter=delimiter)
                        first_row = next(csv_reader, None)

                        if first_row is not None:
                            column_count = len(first_row)
                            data = str(column_count)
                        else:
                            data = '0'
                    else:
                        data = 'delimiter-not-found'
                except Exception as e:
                    data = ''
                    error_string = error_string + str(e)
                    logger.info(error_string)
            else:
                data = 'get-filesize-over-error'

        elif key == 'get-text-delimiter' and extension in AddonAppConfig.text_file_extension:
            if file_size < text_file_maximum_size:
                response = download_file(url, cookies)
                try:
                    content = response.decode('utf-8')
                    object = io.StringIO(content)
                    simple_lines = object.getvalue().splitlines()
                    delimiters = AddonAppConfig.delimiters
                    delimiter_counts = {delimiter: 0 for delimiter in delimiters}

                    first_delimiter = None

                    for line in simple_lines:
                        for delimiter in delimiters:
                            count = line.count(delimiter)
                            if count > 0:
                                delimiter_counts[delimiter] += count
                                if first_delimiter is None:
                                    first_delimiter = delimiter

                    # Determine the most frequent delimiter
                    max_count = max(delimiter_counts.values())
                    most_frequent_delimiters = [
                        delimiter for delimiter, count in delimiter_counts.items() if count == max_count
                    ]

                    if len(most_frequent_delimiters) == 1:
                        # If only one delimiter has the maximum count, use it
                        data = delimiters[most_frequent_delimiters[0]]
                    else:
                        # If there's a tie, use the first delimiter that appeared in the file
                        data = delimiters[first_delimiter]

                except Exception as e:
                    data = ''
                    error_string = error_string + str(e)
                    logger.info(error_string)
            else:
                data = 'get-filesize-over-error'

        elif key == 'get-image-type' and extension in AddonAppConfig.image_file_extension:
            if file_size < image_file_maximum_size:
                response = download_file(url, cookies)
                try:
                    image_type = imghdr.what(None, h=response)
                    if image_type is not None:
                        data = str(image_type)
                    else:
                        data = str('')
                except Exception as e:
                    data = ''
                    error_string = error_string + str(e)
                    logger.info(error_string)
            else:
                data = 'get-filesize-over-error'

        elif key == 'get-image-color-information' and extension in AddonAppConfig.image_file_extension:
            if file_size < image_file_maximum_size:
                response = download_file(url, cookies)
                try:
                    image_data = b''

                    with requests.get(url, cookies=cookies, stream=True) as r:
                        for chunk in r.iter_content(chunk_size=1024 * 1024):
                            if chunk:
                                image_data += chunk
                                try:
                                    image = Image.open(BytesIO(image_data)).convert('RGB')
                                    w, h = image.size
                                    pixels = image.load()

                                    is_black = True
                                    for i in range(w):
                                        for j in range(h):
                                            r, g, b = pixels[i, j]
                                            if r != g or r != b or g != b:
                                                is_black = False
                                                break
                                        if not is_black:
                                            break

                                    if is_black:
                                        data = 'monochrome'
                                    else:
                                        data = 'color'
                                except Exception:
                                    continue
                except Exception as e:
                    data = ''
                    error_string = error_string + str(e)
                    logger.info(error_string)
            else:
                data = 'get-filesize-over-error'

        elif key == 'get-image-text/binary' and (extension in AddonAppConfig.text_file_extension or extension in AddonAppConfig.excel_file_extension or extension in AddonAppConfig.image_file_extension):
            if file_size < any_file_maximum_size:
                response = download_file(url, cookies)
                try:
                    # Download the content directly into a variable
                    with requests.get(url, cookies=cookies, stream=True) as r:
                        file_content = b''
                        for chunk in r.iter_content(chunk_size=1024 * 1024):
                            file_content += chunk

                    # Determine if the content is binary or text
                    if is_binary(file_content, extension):
                        data = 'binary'
                    else:
                        data = 'text'

                except Exception as e:
                    data = ''
                    error_string = error_string + str(e)
                    logger.info(error_string)
            else:
                data = 'get-filesize-over-error'

        elif key == 'get-image-resolution' and extension in AddonAppConfig.image_file_extension:
            if file_size < image_file_maximum_size:
                response = download_file(url, cookies)
                try:
                    Image.MAX_IMAGE_PIXELS = None
                    image = Image.open(BytesIO(response))

                    dpi_info = image.info.get('dpi')
                    if dpi_info:
                        horizontal_dpi = dpi_info[0]

                    else:
                        horizontal_dpi = 'Unknown'

                    data = f'{horizontal_dpi} dpi'
                except Exception as e:
                    data = ''
                    error_string = error_string + str(e)
                    logger.info(error_string)
            else:
                data = 'get-filesize-over-error'

        elif key == 'get-text-character-code' and extension in AddonAppConfig.text_file_extension:
            if file_size < text_file_maximum_size:
                response = download_file(url, cookies)
                try:
                    encoding = chardet.detect(response).get('encoding')
                    data = str(encoding)
                except Exception as e:
                    data = ''
                    error_string = error_string + str(e)
                    logger.info(error_string)
            else:
                data = 'get-filesize-over-error'

        elif key == 'get-image-datasaize' and extension in AddonAppConfig.image_file_extension:
            if file_size < image_file_maximum_size:
                response = download_file(url, cookies)
                try:
                    Image.MAX_IMAGE_PIXELS = None
                    image = Image.open(BytesIO(response))
                    width, height = image.size
                    data = str(width) + ' x ' + str(height)
                except Exception as e:
                    data = ''
                    error_string = error_string + str(e)
                    logger.info(error_string)
            else:
                data = 'get-filesize-over-error'

    except Exception as e:
        data = ''
        error_string = error_string + str(e)
        logger.info(error_string)
    return [{
        'key': key,
        'value': data,
        'error': error_string,
    }]

def is_binary(file_content, extension):
    """Binary file detection.
    Determine by MIME type and content if MIME is not enough.
    """
    ret = False
    m = mimetypes.guess_type('dummy.' + extension)[0]
    if m is None:
        encode = chardet.detect(file_content)['encoding']

        # If encoding is None, consider it binary
        if encode is None:
            ret = True
        elif is_binary08H_character(file_content):
            ret = True
    else:
        if 'office' in m or m.startswith('image') or m.startswith('application/vnd.ms-excel'):
            ret = True
        elif m.startswith('text'):
            ret = False

    return ret

def is_binary08H_character(buf):
    """Binary file detection.
    Check for ASCII code below 08H to determine binary.
    """
    ret = False
    for b in buf:
        if b < 9:
            ret = True
            break
    return ret

def read_dump(dumpname, mode='r'):
    buf = ''
    with open(dumpname, mode) as f:
        buf = f.read()
    return buf

def detect_delimiter(content, delimiters):
    csv_data = io.StringIO(content)
    lines = csv_data.getvalue().splitlines()  # Read all lines
    for line in lines:
        for delimiter in delimiters:
            if delimiter in line:
                return delimiter
    return None

def get_file_size_and_extension(url, cookies):
    try:
        response = requests.head(url, cookies=cookies, stream=True)
        response.raise_for_status()
        content_length = response.headers.get('Content-Length')
        if content_length is not None:
            file_size = int(content_length)
        else:
            raise ValueError('The Content-Length header is missing.')

        content_disposition = response.headers.get('X-Waterbutler-Metadata')
        metadata_json = json.loads(content_disposition)
        file_name = metadata_json.get('attributes', {}).get('name')
        extension = file_name.split('.')[-1]

        return file_size, extension
    except requests.RequestException as e:
        logger.info(str(e))
        return None, None

def download_file(url, cookies):
    response = requests.get(url, cookies=cookies).content
    return response

def suggestion_file_data_number(key, filepath, node):
    parts = filepath.split('/')
    is_dir = parts[0] == 'dir'
    if is_dir:
        value = 'files/{}'.format(filepath)
    else:
        provider = parts[0]
        path = '/'.join(parts[1:])
        try:
            file_node = BaseFileNode.resolve_class(provider, BaseFileNode.FILE).get_or_create(node, path)
        except UnableToResolveFileClass:
            raise HTTPError(http_status.HTTP_404_NOT_FOUND)
        guid = file_node.get_guid(create=True)
        guid.referent.save()
        value = guid._id
    return [{
        'key': key,
        'value': value,
    }]


def suggestion_ror(key, keyword):
    response = requests.get(
        ROR_URL,
        params={
            'query': keyword,
        }
    )
    response.raise_for_status()
    res = []
    for item in response.json()['items']:
        names = item.get('names', [])
        display_name = next((n['value'] for n in names if 'ror_display' in n.get('types', [])),
                           next((n['value'] for n in names if n.get('lang') == 'en'), ''))
        name_ja = next((n['value'] for n in names if n.get('lang') == 'ja' and 'label' in n.get('types', [])),
                      next((n['value'] for n in names if n.get('lang') == 'ja'), display_name))
        item['name'] = display_name
        res.append({
            'key': key,
            'value': {
                **item,
                'name-ja': name_ja,
            }
        })
    return res


def suggestion_asset(key, keyword, node):
    addon = node.get_addon(SHORT_NAME)
    assets = addon.get_metadata_assets()
    res = []
    for asset in assets:
        key_target = asset.get(key[6:], '').lower()
        if len(key_target) > 0 and keyword in key_target:
            res.append({
                'key': key,
                'value': asset,
            })
    return res
