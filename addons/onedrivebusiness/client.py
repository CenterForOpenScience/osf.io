# -*- coding: utf-8 -*-
import io
import logging

from framework.exceptions import HTTPError

from addons.onedrive import settings  # using base settings
from addons.onedrive.client import OneDriveClient
from addons.onedrivebusiness import settings as own_settings

from openpyxl import Workbook, load_workbook


logger = logging.getLogger(__name__)


class OneDriveBusinessClient(OneDriveClient):

    def __init__(self, access_token=None, drive_id=None):
        super(OneDriveBusinessClient, self).__init__(access_token=access_token, drive_id=drive_id)

    def files(self, folder_id=None):
        """Get list of files of the folder with id ``folder_id``

        API Docs:  https://docs.microsoft.com/en-us/graph/api/driveitem-list-children

        :param str folder_id: the id of the parent folder. defaults to ``None``
        :rtype: list
        :return: a list of metadata objects representing the child files of ``folder_id``
        """

        if folder_id is None:
            raise ValueError('Root folder must be specified')
        url = self._build_url(self._drive_url, 'items', folder_id, 'children')

        resp = self._make_request(
            'GET',
            url,
            expects=(200, ),
            throws=HTTPError(401)
        )
        data = resp.json()
        logger.debug('Response: {}'.format(data))
        res = data['value']

        next_url = data.get('children@odata.nextLink', None)
        while next_url is not None:
            next_resp = self._make_request(
                'GET',
                next_url,
                expects=(200, ),
                throws=HTTPError(401)
            )
            next_data = next_resp.json()
            res.extend(next_data['value'])
            next_url = next_data.get('@odata.nextLink', None)

        return res

    def create_file(self, parent_folder_id, name, content):
        """Upload file in ``parent_folder_id`` with ``name`` and ``content``

        API Docs:  https://docs.microsoft.com/en-us/graph/api/driveitem-put-content

        :param str parent_folder_id: the id of the parent folder.
        :param str name: the name of new file.
        :param bytes content: the content of new file.
        :rtype: dict
        :return: a metadata object representing the new file
        """

        if parent_folder_id is None:
            raise ValueError('Root folder must be specified')
        url = self._build_url(self._drive_url, 'items',
                              parent_folder_id + ':', name + ':', 'content')

        res = self._make_request(
            'PUT',
            url,
            data=content,
            expects=(200, 201),
            throws=HTTPError(401)
        )
        return res.json()

    def get_file_content(self, file_id):
        """Get content of ``file_id``

        API Docs:  https://docs.microsoft.com/en-us/graph/api/driveitem-get-content

        :param str file_id: the id of the file.
        :rtype: dict
        :return: a bytes array of the content of file
        """

        url = self._build_url(self._drive_url, 'items',
                              file_id, 'content')

        res = self._make_request(
            'GET',
            url,
            expects=(200,),
            throws=HTTPError(401)
        )
        return res.content

    def create_folder(self, parent_folder_id, name):
        """Create new folder in ``parent_folder_id`` with ``name``

        API Docs:  https://docs.microsoft.com/en-us/graph/api/driveitem-post-children

        :param str parent_folder_id: the id of the parent folder.
        :param str name: the name of new folder.
        :rtype: dict
        :return: a metadata object representing the new folder
        """

        url = self._build_url(self._drive_url, 'items', parent_folder_id, 'children')
        res = self._make_request(
            'POST',
            url,
            json={
                'name': name,
                'folder': {},
                '@microsoft.graph.conflictBehavior': 'fail'
            },
            expects=(200, 201),
            throws=HTTPError(401)
        )
        return res.json()

    def rename_folder(self, folder_id, name):
        """Rename a folder specified by ``folder_id`` with ``name``

        API Docs:  https://docs.microsoft.com/en-us/graph/api/driveitem-update

        :param str folder_id: the id of the folder.
        :param str name: the name of the folder.
        :rtype: dict
        :return: a metadata object representing the folder
        """

        url = self._build_url(self._drive_url, 'items', folder_id)
        res = self._make_request(
            'PATCH',
            url,
            json={
                'name': name,
            },
            expects=(200,),
            throws=HTTPError(401)
        )
        return res.json()

    def get_user(self, msaccount):
        url = self._build_url(settings.ONEDRIVE_API_URL, 'users', msaccount)
        res = self._make_request(
            'GET',
            url,
            expects=(200,),
            throws=HTTPError(401)
        )
        return res.json()

    def get_drive(self, me=False):
        if me:
            url = self._build_url(settings.ONEDRIVE_API_URL, 'me', 'drive')
        else:
            url = self._build_url(settings.ONEDRIVE_API_URL, 'drive')
        res = self._make_request(
            'GET',
            url,
            expects=(200,),
            throws=HTTPError(401)
        )
        return res.json()

    def get_permissions(self, item_id):
        url = self._build_url(self._drive_url, 'items', item_id, 'permissions')
        res = self._make_request(
            'GET',
            url,
            expects=(200,),
            throws=HTTPError(401)
        )
        return res.json()

    def invite_user(self, item_id, recipient_email, roles):
        url = self._build_url(self._drive_url, 'items', item_id, 'invite')
        res = self._make_request(
            'POST',
            url,
            json={
                'recipients': [{'email': recipient_email}],
                'requireSignIn': True,
                'sendInvitation': False,
                'roles': roles,
            },
            expects=(200,),
            throws=HTTPError(401)
        )
        return res.json()

    def remove_permission(self, item_id, permission_id):
        url = self._build_url(self._drive_url, 'items', item_id, 'permissions', permission_id)
        self._make_request(
            'DELETE',
            url,
            expects=(200, 204),
            throws=HTTPError(401)
        )

    def update_permission(self, item_id, permission_id, roles):
        url = self._build_url(self._drive_url, 'items', item_id, 'permissions', permission_id)
        res = self._make_request(
            'PATCH',
            url,
            json={
                'roles': roles,
            },
            expects=(200,),
            throws=HTTPError(401)
        )
        return res.json()


class UserListClient:

    def __init__(self, client, folder_id, filename=None, sheet_name=None):
        self.client = client
        self.folder_id = folder_id
        self.filename = filename if filename is not None else own_settings.TEAM_MEMBER_LIST_FILENAME
        self.sheet_name = sheet_name if sheet_name is not None else own_settings.TEAM_MEMBER_LIST_SHEETNAME
        self.workbook = None

    def get_workbook_sheet(self):
        wb = self._load_workbook()
        if wb is None:
            wb = self._prepare_workbook()
        return wb[self.sheet_name]

    def _load_workbook(self):
        files = [f for f in self.client.files(self.folder_id) if f['name'] == self.filename]
        if len(files) == 0:
            return None
        content = io.BytesIO(self.client.get_file_content(files[0]['id']))
        return load_workbook(filename=content)

    def _prepare_workbook(self):
        wb = Workbook()
        worksheet = wb.active
        worksheet.title = self.sheet_name
        worksheet['A1'] = 'ePPN'
        worksheet['B1'] = 'MicrosoftAccount'
        worksheet['A2'] = '#ここにePPNを記述'
        worksheet['B2'] = '#ここにMicrosoftアカウントを記述'
        buf = io.BytesIO()
        wb.save(filename=buf)
        self.client.create_file(self.folder_id, self.filename, buf.getvalue())
        return wb
